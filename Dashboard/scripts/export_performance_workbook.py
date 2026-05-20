from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ROI_WORKBOOK_PATH = PROJECT_ROOT / "Data" / "Client Performance Tracking (3).xlsx"
META_WORKBOOK_PATH = PROJECT_ROOT / "Data" / "Ads ROI - Sheet.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "Data" / "performance-dashboard.json"

EXCLUDED_ROI_SHEETS = {
    "Overview",
    "Template",
    "Copy of Template",
    "Copy of Template 1",
}

DISCOVERY_LABELS = {"followers", "new leads", "leads"}
RETARGETING_LABELS = {"retargeting"}

META_MONTH_PATTERNS = [
    ("january", 1),
    ("jan", 1),
    ("february", 2),
    ("feb", 2),
    ("march", 3),
    ("mar", 3),
    ("april", 4),
    ("apr", 4),
    ("may", 5),
    ("june", 6),
    ("jun", 6),
    ("july", 7),
    ("jul", 7),
    ("august", 8),
    ("aug", 8),
    ("september", 9),
    ("sept", 9),
    ("sep", 9),
    ("october", 10),
    ("oct", 10),
    ("november", 11),
    ("nov", 11),
    ("december", 12),
    ("dec", 12),
]

META_COMMENT_KEYS = (
    "comments",
    "comment",
    "notes",
    "recommended_action",
    "todos",
    "todo",
)


def parse_meta_sheet_config(sheet_name: str) -> dict[str, object] | None:
    text = clean_text(sheet_name).lower()
    year_match = re.search(r"\b(20\d{2})\b", text)
    if not year_match:
        return None

    month = None
    for token, month_number in META_MONTH_PATTERNS:
        if re.search(r"\b" + re.escape(token) + r"\b", text):
            month = month_number
            break

    if month is None:
        return None

    return {"name": sheet_name, "year": int(year_match.group(1)), "month": month}


def find_meta_header_row(sheet) -> tuple[int | None, list[str]]:
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        headers = [normalize_header(value) for value in row]
        if "campaign_type" in headers and ("spend" in headers or "campaign_spend" in headers):
            return row_index, headers
    return None, []


def build_meta_row_map(headers: list[str], row: tuple[object, ...]) -> dict[str, object]:
    row_map: dict[str, object] = {}
    seen: dict[str, int] = {}
    last_header: str | None = None
    for index, header in enumerate(headers):
        if index >= len(row):
            continue
        normalized_header = header
        if not normalized_header:
            if not last_header:
                continue
            normalized_header = last_header
        seen[normalized_header] = seen.get(normalized_header, 0) + 1
        key = normalized_header if seen[normalized_header] == 1 else f"{normalized_header}_{seen[normalized_header]}"
        row_map[key] = row[index]
        last_header = normalized_header
    return row_map

MANUAL_META_CLIENTS = {
    "asheville river cabins": ("asheville-river-cabins", "Asheville River Cabins"),
    "asheville river cabins ads": ("asheville-river-cabins", "Asheville River Cabins"),
    "away 2pa": ("away2pa", "Away2PA"),
    "away2pa": ("away2pa", "Away2PA"),
    "awayframes ad account": ("awayframes", "Awayframes"),
    "big moon ranch": ("big-moon-ranch", "Big Moon Ranch"),
    "cohost": ("the-cohost-company", "The Cohost Company"),
    "the cohost company ads": ("the-cohost-company", "The Cohost Company"),
    "cowan creek lodges ads": ("cowan-creek-lodges", "Cowan Creek Lodges"),
    "creekside": ("creekside", "Creekside"),
    "dwell luxury rentals ads": ("dwell-luxury-rentals", "Dwell Luxury Rentals"),
    "edenwood nc": ("edenwood-nc", "Edenwood NC"),
    "endless stays": ("endless-stays", "Endless Stays"),
    "evergreen": ("evergreen-cabins", "Evergreen Cabins"),
    "evergreen cabins ads": ("evergreen-cabins", "Evergreen Cabins"),
    "flohom": ("flohom", "FLOHOM"),
    "flohom ads": ("flohom", "FLOHOM"),
    "green springs inn": ("green-springs-inn", "Green Springs Inn"),
    "hiawassee glamping": ("hiawassee-glamping", "Hiawassee Glamping"),
    "hillside amble": ("hillside-amble", "Hillside Amble"),
    "home base": ("home-base", "Home Base"),
    "home base bnbs": ("home-base", "Home Base"),
    "inspired retreats": ("inspired-retreats", "Inspired Retreats"),
    "myrinn": ("myrinn", "Myrinn"),
    "myrinn ad account 2 0": ("myrinn", "Myrinn"),
    "nature nooks": ("nature-nooks", "Nature Nooks"),
    "paradise pointe": ("paradise-pointe", "Paradise Pointe"),
    "paradise pointe ads": ("paradise-pointe", "Paradise Pointe"),
    "parker reserve": ("parker-reserve", "Parker Reserve"),
    "pine valley cabins georgia ad account": ("pine-valley-cabins", "Pine Valley Cabins"),
    "reflections resort": ("reflections-resorts", "Reflections Resorts"),
    "reflections resort ad": ("reflections-resorts", "Reflections Resorts"),
    "roundhouse residences ads": ("roundhouse-resort-spa", "Roundhouse Resort & Spa"),
    "starlight hotsprings": ("starlight-haven-hot-springs", "Starlight Haven Hot Springs"),
    "starlight haven hot springs": ("starlight-haven-hot-springs", "Starlight Haven Hot Springs"),
    "starlight weiss lake": ("starlight-haven-weiss-lake", "Starlight Haven Weiss Lake"),
    "starlight haven weiss lake": ("starlight-haven-weiss-lake", "Starlight Haven Weiss Lake"),
    "stay different": ("stay-different", "Stay Different"),
    "stay luxe ads": ("stayluxe", "StayLuxe"),
    "stay on 30a ads": ("stay-on-30a", "Stay on 30a"),
    "stay saluda": ("stay-saluda", "Stay Saluda"),
    "stay saluda ads": ("stay-saluda", "Stay Saluda"),
    "stay with branch": ("stay-with-branch", "Stay With Branch"),
    "bison ridge": ("bison-ridge-retreat", "Bison Ridge Retreat"),
    "bison ridge retreat": ("bison-ridge-retreat", "Bison Ridge Retreat"),
    "three suns": ("three-suns", "Three Suns"),
    "treetop": ("treetop-escapes", "Treetop Escapes"),
    "treetop escapes": ("treetop-escapes", "Treetop Escapes"),
    "treetop escapes ads": ("treetop-escapes", "Treetop Escapes"),
    "wanderin star farms": ("wanderin-star-farms", "Wanderin Star Farms"),
    "wanderin star farms new ad account": ("wanderin-star-farms", "Wanderin Star Farms"),
    "wauhatchie woodlands": ("wauhatchie-woodlands", "Wauhatchie Woodlands"),
    "wauhatchie woodlands timberroot east coast": ("wauhatchie-woodlands", "Wauhatchie Woodlands"),
    "yosemite meta ads": ("yosemite-dream-stays", "Yosemite Dream Stays"),
}

MANUAL_ROI_CLIENTS = {
    "bison ridge retreat": ("bison-ridge-retreat", "Bison Ridge Retreat"),
    "three suns cabins": ("three-suns", "Three Suns"),
}


def slugify(value: str) -> str:
    return re.sub(r"(^-|-$)", "", re.sub(r"[^a-z0-9]+", "-", value.lower()))


def clean_text(value: object) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def normalize_name(value: object) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return " ".join(text.split())


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = (
        text.replace("🎉", "")
        .replace("📈", "")
        .replace("👁️", "")
        .replace("/", " ")
        .replace("%", " pct ")
        .replace("#", " num ")
        .replace("$", " ")
    )
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def clean_number(value: object):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text.upper() in {"N/A", "#DIV/0!"}:
        return None

    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace(",", "").replace("$", "").replace("%", "")
    try:
        number = float(text)
        return -number if negative else number
    except ValueError:
        return None


def numeric(value: object) -> float:
    return clean_number(value) or 0.0


def percent_text(value: object):
    if value is None or value == "":
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None

    number = clean_number(value)
    if number is None:
        return None
    if abs(number) <= 1:
        return f"{round(number * 100)}%"
    return f"{round(number)}%"


def get_value(row_map: dict[str, object], *keys: str):
    for key in keys:
        if key in row_map:
            return row_map[key]
    return None


def get_campaign_spend(row_map: dict[str, object]):
    """
    Most campaign rows carry the usable spend in the second spend column, but
    January and December only have a single spend column in the source workbook.
    Use the first spend column for those months and the split spend column for
    the rest.
    """
    timeline = get_value(row_map, "timeline", "month")
    month = getattr(timeline, "month", None)
    if month in {1, 12}:
        return get_value(
            row_map,
            "spend",
            "campaign_spend",
            "ad_spend",
            "spend_2",
            "campaign_spend_2",
            "ad_spend_2",
        )
    return get_value(
        row_map,
        "spend_2",
        "campaign_spend_2",
        "ad_spend_2",
        "spend",
        "campaign_spend",
        "ad_spend",
    )


def collect_comments(row_map: dict[str, object], *comment_keys: str) -> str:
    values = []
    for key in comment_keys:
        value = get_value(row_map, key)
        if value in (None, "", True, False):
            continue
        if isinstance(value, str):
            text = clean_text(value)
            if text:
                values.append(text)
    return " | ".join(dict.fromkeys(values))


def find_roi_header_row(sheet) -> tuple[int | None, list[str]]:
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        headers = [normalize_header(value) for value in row]
        if "timeline" in headers or "month" in headers:
            return row_index, headers
    return None, []


def build_roi_row(row_map: dict[str, object]):
    timeline = get_value(row_map, "timeline", "month")
    if not isinstance(timeline, datetime):
        return None

    return {
        "year": timeline.year,
        "month": timeline.month,
        "total_views": clean_number(get_value(row_map, "total_views")) or 0,
        "total_view_growth": percent_text(get_value(row_map, "total_view_growth")),
        "ig_views": clean_number(get_value(row_map, "ig_views")) or 0,
        "fb_views": clean_number(get_value(row_map, "fb_views")) or 0,
        "tiktok_views": clean_number(get_value(row_map, "tiktok_views")) or 0,
        "ig_followers": clean_number(get_value(row_map, "ig_followers")) or 0,
        "fb_followers": clean_number(get_value(row_map, "fb_followers")) or 0,
        "tiktok_followers": clean_number(get_value(row_map, "tiktok_followers")) or 0,
        "ttl_followers": clean_number(get_value(row_map, "ttl_followers")) or 0,
        "follower_growth_pct": clean_number(get_value(row_map, "follower_growth")) or 0,
        "website_traffic": clean_number(get_value(row_map, "website_traffic")) or 0,
        "ad_spend": clean_number(get_value(row_map, "ad_spend")) or 0,
        "cost_per_follower": clean_number(get_value(row_map, "cost_per_follower")),
        "cost_per_lead": clean_number(get_value(row_map, "cost_per_lead")),
        "cost_per_booking": clean_number(get_value(row_map, "cost_per_booking")),
        "new_leads": clean_number(get_value(row_map, "new_leads")) or 0,
        "ttl_leads": clean_number(get_value(row_map, "ttl_leads")) or 0,
        "lead_growth_pct": clean_number(get_value(row_map, "lead_growth")) or 0,
        "total_booking_revenue": clean_number(get_value(row_map, "total_booking_revenue")) or 0,
        "direct_booking_revenue": clean_number(get_value(row_map, "direct_booking_revenue")) or 0,
        "direct_booking_split_pct": clean_number(get_value(row_map, "direct_booking_split")) or 0,
        "ly_total_booking_revenue": clean_number(get_value(row_map, "ly_total_booking_revenue")) or 0,
        "ly_direct_booking_revenue": clean_number(get_value(row_map, "ly_direct_booking_revenue")) or 0,
        "ly_direct_booking_split_pct": clean_number(get_value(row_map, "ly_direct_booking_split")) or 0,
        "notes": str(get_value(row_map, "notes_insights") or "").strip(),
    }


@dataclass
class CampaignAccumulator:
    rows: list[dict[str, object]] = field(default_factory=list)

    def add_row(self, row: dict[str, object]) -> None:
        self.rows.append(row)

    def as_payload(self, year: int, month: int, campaign_type: str, avg_booking_value: float | None, comments: str) -> dict[str, object] | None:
        if not self.rows:
            return None

        if len(self.rows) == 1:
            source = self.rows[0]
            return {
                "year": year,
                "month": month,
                "campaign_type": campaign_type,
                "spend": clean_number(source["spend"]),
                "impressions": clean_number(source["impressions"]),
                "profile_visits": clean_number(source["profile_visits"]),
                "cost_per_visit": clean_number(source["cost_per_visit"]),
                "leads_followers": clean_number(source["leads_followers"]),
                "cost_per_lead_follower": clean_number(source["cost_per_lead_follower"]),
                "ig_bio_leads": clean_number(source["ig_bio_leads"]),
                "bookings_email_matched": clean_number(source["bookings_email"]),
                "bookings_fb_events": clean_number(source["bookings_fb"]),
                "cost_per_booking": clean_number(source["cost_per_booking"]),
                "avg_booking_value": avg_booking_value or clean_number(source["avg_booking_value"]),
                "pct_avg_booking_value": clean_number(source["pct_avg_booking_value"]),
                "revenue": clean_number(source["revenue"]),
                "roas": clean_number(source["roas"]),
                "blended_roas": clean_number(source["blended_roas"]),
                "comments": comments or clean_text(source.get("comments")),
            }

        spend = sum(numeric(row["spend"]) for row in self.rows)
        impressions = sum(numeric(row["impressions"]) for row in self.rows)
        profile_visits = sum(numeric(row["profile_visits"]) for row in self.rows)
        leads_followers = sum(numeric(row["leads_followers"]) for row in self.rows)
        ig_bio_leads = sum(numeric(row["ig_bio_leads"]) for row in self.rows)
        bookings_email = sum(numeric(row["bookings_email"]) for row in self.rows)
        bookings_fb = sum(numeric(row["bookings_fb"]) for row in self.rows)
        total_bookings = bookings_email + bookings_fb
        revenue = sum(numeric(row["revenue"]) for row in self.rows)

        weighted_blended_total = 0.0
        for row in self.rows:
            row_spend = numeric(row["spend"])
            if row_spend <= 0:
                continue
            weighted_blended_total += row_spend * (clean_number(row["blended_roas"]) or clean_number(row["roas"]) or 0.0)

        cost_per_visit = (spend / profile_visits) if profile_visits > 0 else None
        cost_per_lead_follower = (spend / leads_followers) if leads_followers > 0 else None
        cost_per_booking = (spend / total_bookings) if total_bookings > 0 else None
        pct_avg_booking_value = (cost_per_booking / avg_booking_value) if cost_per_booking is not None and avg_booking_value else None
        roas = (revenue / spend) if spend > 0 and revenue > 0 else 0
        blended_roas = (weighted_blended_total / spend) if spend > 0 and weighted_blended_total > 0 else roas

        return {
            "year": year,
            "month": month,
            "campaign_type": campaign_type,
            "spend": round(spend, 2),
            "impressions": round(impressions),
            "profile_visits": round(profile_visits),
            "cost_per_visit": cost_per_visit,
            "leads_followers": round(leads_followers),
            "cost_per_lead_follower": cost_per_lead_follower,
            "ig_bio_leads": round(ig_bio_leads),
            "bookings_email_matched": round(bookings_email),
            "bookings_fb_events": round(bookings_fb),
            "cost_per_booking": cost_per_booking,
            "avg_booking_value": avg_booking_value,
            "pct_avg_booking_value": pct_avg_booking_value,
            "revenue": round(revenue, 2),
            "roas": roas,
            "blended_roas": blended_roas,
            "comments": comments,
        }


@dataclass
class MetaClientMonth:
    slug: str
    name: str
    year: int
    month: int
    avg_booking_value: float | None = None
    discovery: CampaignAccumulator = field(default_factory=CampaignAccumulator)
    retargeting: CampaignAccumulator = field(default_factory=CampaignAccumulator)
    comments: list[str] = field(default_factory=list)


def resolve_meta_client(raw_name: object) -> tuple[str, str]:
    normalized = normalize_name(raw_name)
    if normalized in MANUAL_META_CLIENTS:
        return MANUAL_META_CLIENTS[normalized]

    return slugify(normalized), clean_text(raw_name)


def classify_campaign(value: object) -> str | None:
    label = normalize_name(value)
    if label in DISCOVERY_LABELS:
        return "Discovery"
    if label in RETARGETING_LABELS:
        return "Retargeting"
    return None


def export_roi_workbook():
    workbook = load_workbook(ROI_WORKBOOK_PATH, data_only=True)
    clients = []
    rows_by_client_slug = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name in EXCLUDED_ROI_SHEETS:
            continue

        sheet = workbook[sheet_name]
        header_row_index, headers = find_roi_header_row(sheet)
        if not header_row_index:
            continue
        rows = []

        for raw_row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not any(value not in (None, "") for value in raw_row):
                continue
            row_map = {headers[index]: raw_row[index] for index in range(min(len(headers), len(raw_row)))}
            normalized = build_roi_row(row_map)
            if normalized:
                rows.append(normalized)

        if not rows:
            continue

        normalized_name = normalize_name(sheet_name)
        if normalized_name in MANUAL_ROI_CLIENTS:
            slug, name = MANUAL_ROI_CLIENTS[normalized_name]
        else:
            slug, name = slugify(sheet_name), sheet_name
        rows.sort(key=lambda row: (row["year"], row["month"]))
        clients.append({"slug": slug, "name": name})
        rows_by_client_slug[slug] = rows

    clients.sort(key=lambda client: client["name"].lower())
    return clients, rows_by_client_slug


def export_meta_workbook():
    workbook = load_workbook(META_WORKBOOK_PATH, data_only=True)
    clients_by_slug = {}
    meta_rows_by_client_slug: dict[str, list[dict[str, object]]] = {}

    sheet_configs = []
    for sheet_name in workbook.sheetnames:
        sheet_config = parse_meta_sheet_config(sheet_name)
        if sheet_config:
            sheet_configs.append(sheet_config)

    sheet_configs.sort(key=lambda config: (config["year"], config["month"], config["name"].lower()))

    for sheet_config in sheet_configs:
        sheet = workbook[sheet_config["name"]]
        header_row_index, headers = find_meta_header_row(sheet)
        if not header_row_index:
            continue

        current: MetaClientMonth | None = None
        month_blocks: list[MetaClientMonth] = []

        for row in sheet.iter_rows(min_row=header_row_index + 1, max_col=sheet.max_column, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue

            row_map = build_meta_row_map(headers, row)
            raw_name = clean_text(row[0])
            if raw_name:
                if current is not None:
                    month_blocks.append(current)
                slug, name = resolve_meta_client(raw_name)
                current = MetaClientMonth(
                    slug=slug,
                    name=name,
                    year=sheet_config["year"],
                    month=sheet_config["month"],
                )

            if current is None:
                continue

            avg_booking_value = clean_number(
                get_value(row_map, "avg_booking_value", "average_booking_value")
            )
            if avg_booking_value is not None:
                current.avg_booking_value = avg_booking_value

            campaign_type = classify_campaign(get_value(row_map, "campaign_type"))
            if campaign_type is None:
                continue

            comment = collect_comments(row_map, *META_COMMENT_KEYS)
            if comment:
                current.comments.append(comment)

            payload = {
                "spend": get_campaign_spend(row_map),
                "impressions": get_value(row_map, "impressions"),
                "profile_visits": get_value(
                    row_map,
                    "profile_website_visits_link_clicks",
                    "profile_website_visits",
                    "profile_website_visits_visits",
                ),
                "cost_per_visit": get_value(row_map, "cost_per_visit"),
                "leads_followers": get_value(row_map, "leads_followers"),
                "cost_per_lead_follower": get_value(
                    row_map,
                    "cost_per_lead_follower",
                    "cost_lead_or_follower",
                    "cost_per_follower",
                    "cost_per_lead",
                ),
                "ig_bio_leads": get_value(row_map, "instagram_bio_leads", "ig_bio_leads"),
                "bookings_email": get_value(
                    row_map,
                    "bookings_email_cross_matched_code_used",
                    "bookings_email_matched",
                    "bookings_email",
                ),
                "bookings_fb": get_value(
                    row_map,
                    "bookings_facebook_events",
                    "bookings_fb_events",
                    "bookings_fb",
                ),
                "cost_per_booking": get_value(row_map, "cost_per_booking"),
                "avg_booking_value": get_value(row_map, "avg_booking_value", "average_booking_value"),
                "pct_avg_booking_value": get_value(
                    row_map,
                    "pct_avg_booking_value",
                    "pct_of_avg_booking_value_spent_on_getting_a_booking",
                ),
                # Meta ads revenue should come from the ad-specific revenue columns,
                # not the ROI workbook's total direct booking revenue field.
                "revenue": get_value(
                    row_map,
                    "ad_revenue",
                    "revenue",
                    "ads_rev_value",
                ),
                "roas": get_value(row_map, "roas_2", "roas"),
                "blended_roas": get_value(row_map, "blended_roas"),
                "comments": comment,
            }

            if campaign_type == "Discovery":
                current.discovery.add_row(payload)
            else:
                current.retargeting.add_row(payload)

            clients_by_slug[current.slug] = {"slug": current.slug, "name": current.name}

        if current is not None:
            month_blocks.append(current)

        for block in month_blocks:
            bucket = meta_rows_by_client_slug.setdefault(block.slug, [])
            comments = " | ".join(dict.fromkeys(block.comments))
            discovery_row = block.discovery.as_payload(block.year, block.month, "Discovery", block.avg_booking_value, comments)
            retargeting_row = block.retargeting.as_payload(block.year, block.month, "Retargeting", block.avg_booking_value, comments)
            if discovery_row:
                discovery_row["_month_key"] = (block.year, block.month, "Discovery")
                bucket.append(discovery_row)
            if retargeting_row:
                retargeting_row["_month_key"] = (block.year, block.month, "Retargeting")
                bucket.append(retargeting_row)

    for rows in meta_rows_by_client_slug.values():
        rows.sort(key=lambda row: (row["year"], row["month"], row["campaign_type"]))
        for row in rows:
            row.pop("_month_key", None)

    clients = sorted(clients_by_slug.values(), key=lambda client: client["name"].lower())
    return clients, meta_rows_by_client_slug


def export_workbook():
    roi_clients, rows_by_client_slug = export_roi_workbook()
    meta_clients, meta_rows_by_client_slug = export_meta_workbook()

    client_map = {client["slug"]: client for client in roi_clients}
    for client in meta_clients:
        client_map.setdefault(client["slug"], client)

    payload = {
        "clients": sorted(client_map.values(), key=lambda client: client["name"].lower()),
        "rowsByClientSlug": rows_by_client_slug,
        "metaRowsByClientSlug": meta_rows_by_client_slug,
    }

    OUTPUT_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    export_workbook()
